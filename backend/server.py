from fastapi import FastAPI, APIRouter, HTTPException, Header, Depends, Request, Query
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo.errors import DuplicateKeyError
import os
import io
import logging
import re
import secrets
import uuid
import bcrypt
import jwt
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Literal
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')
ADMIN_TOKEN = os.environ.get('ADMIN_TOKEN', 'mbi-nonton-2026-admin')
JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALGO = "HS256"
TOKEN_TTL_HOURS = 12
ROLE_LABELS = {"superadmin": "Super Admin", "admin": "Admin", "checkin": "Petugas Check-in"}

app = FastAPI()
api_router = APIRouter(prefix="/api")

# ---------------- Event configuration ----------------
REFERENCE_COST = 60000  # biaya pengadaan rata-rata per orang (acuan dana sukarela)
HOLD_MINUTES = 15  # unpaid orders auto-release seats after this
MAX_SEATS_PER_ORDER = 6  # batas kursi per pesanan
MAX_PROOF_BYTES = 8_000_000  # ~6MB gambar (anti-DoS upload publik)
LOGIN_MAX_FAILS = 5          # gagal login sebelum dikunci
LOGIN_LOCK_MINUTES = 15      # durasi kunci setelah terlalu banyak gagal
EVENT_TITLE = 'Nonton Bersama Film Dokumenter "Y.A. MNS. Ashin Jinarakkhita: Jejak Langkah Sang Pelopor di Nusantara"'
EVENT_DATE = "Minggu, 13 September 2026"
EVENT_LOCATION = "CGV Grand Batam"

SESSIONS = [
    {"id": 1, "name": "Sesi 1", "time": "09.30–11.30 WIB"},
    {"id": 2, "name": "Sesi 2", "time": "12.00–14.00 WIB"},
    {"id": 3, "name": "Sesi 3", "time": "14.30–16.30 WIB"},
    {"id": 4, "name": "Sesi 4", "time": "17.00–19.00 WIB"},
]

def _rng(a: int, b: int):
    """Descending seat numbers a..b (visual left-to-right, no 1 at right)."""
    return list(range(a, b - 1, -1))


# Denah asli CINEMA 4 (REGULER) CGV Grand Batam — nomor 1 di kanan.
# Blocks = kelompok kursi dipisah gang (aisle).
SEAT_LAYOUT = [
    {"row": "M", "blocks": [[21, 20], _rng(18, 6), [4, 3, 2, 1]]},
    {"row": "L", "blocks": [[21, 20], _rng(18, 6), [4, 3, 2, 1]]},
    {"row": "K", "blocks": [[21, 20], [18, 17, 16, 15, 14, 13, 12, 11, 10, 8, 7, 6], [4, 3]]},
    {"row": "J", "blocks": [[21, 20], _rng(18, 6), [4, 3]]},
    {"row": "H", "blocks": [[21, 20], _rng(18, 6), [4, 3]]},
    {"row": "G", "blocks": [[21, 20], _rng(18, 6), [4, 3]]},
    {"row": "F", "blocks": [[21, 20], _rng(18, 6), [4, 3]]},
    {"row": "E", "blocks": [[21, 20], _rng(18, 6), [4, 3]]},
    {"row": "D", "blocks": [[21, 20], _rng(18, 6), [4, 3]]},
    {"row": "C", "blocks": [[21, 20], _rng(18, 6), [4, 3]]},
    {"row": "B", "blocks": [[21, 20], _rng(16, 7), [4, 3, 2, 1]]},
    {"row": "A", "blocks": [[20, 19, 18, 17], _rng(16, 5), [4, 3, 2, 1]]},
]

# Kursi couple (pink) — wajib dibeli sepasang
COUPLE_PAIRS = [
    ("B16", "B15"), ("B14", "B13"), ("B12", "B11"), ("B10", "B9"), ("B8", "B7"),
    ("A20", "A19"), ("A18", "A17"), ("A16", "A15"), ("A14", "A13"), ("A12", "A11"),
    ("A10", "A9"), ("A8", "A7"), ("A6", "A5"), ("A4", "A3"), ("A2", "A1"),
]
COUPLE_PARTNER = {}
for _a, _b in COUPLE_PAIRS:
    COUPLE_PARTNER[_a] = _b
    COUPLE_PARTNER[_b] = _a

ALL_SEAT_LABELS = {f"{r['row']}{n}" for r in SEAT_LAYOUT for blk in r["blocks"] for n in blk}

# Kursi khusus operator — tidak bisa dipesan di semua sesi
RESERVED_SEATS = {"A11", "A12"}
# Kursi disabilitas — hanya bisa dibeli di lokasi (walk-in), tidak bisa online
DISABILITY_SEATS = {"K10", "K8"}
SEATS_PER_SESSION = len(ALL_SEAT_LABELS) - len(RESERVED_SEATS)  # 207

TRANSFER_INFO = {
    "bank": "BCA",
    "account_number": "061 518 3381",
    "account_name": "PD Majelis Buddhayana Indonesia Prov Kepri",
    "short_name": "PD MBI Kepri",
}


# ---------------- Models ----------------
class OrderCreate(BaseModel):
    name: str
    phone: str
    session_id: int
    seats: List[str]
    payment_method: Literal["qris", "transfer"]
    amount: int = 0  # dana sukarela (total, Rp)


class ProofUpload(BaseModel):
    proof_image: str  # base64 data URL


class WalkinCreate(BaseModel):
    name: str
    phone: Optional[str] = ""
    session_id: int
    seats: List[str]
    payment_method: Literal["cash", "qris", "transfer"]
    amount: int = 0  # dana sukarela (total, Rp)
    proof_image: Optional[str] = None  # wajib untuk qris/transfer (base64 data URL)
    location: Optional[str] = ""  # lokasi penjualan (wajib untuk walk-in)


class AdminLogin(BaseModel):
    username: str
    password: str


class UserCreate(BaseModel):
    username: str
    password: str
    name: Optional[str] = ""
    role: Literal["superadmin", "admin", "checkin"] = "checkin"
    can_delete: bool = False


class UserPermission(BaseModel):
    can_delete: bool


class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[Literal["superadmin", "admin", "checkin"]] = None
    can_delete: Optional[bool] = None


class PasswordReset(BaseModel):
    new_password: str


class VerifyPayload(BaseModel):
    amount: Optional[int] = None  # nominal aktual yang masuk (dari slip TT), opsional


class VIPCreate(BaseModel):
    name: str
    session_id: int
    seats: List[str]
    note: Optional[str] = ""


class SetActiveSession(BaseModel):
    session_id: int


class SessionToggle(BaseModel):
    session_id: int
    open: bool
    target: Literal["public", "walkin"] = "public"


class SetComingSoon(BaseModel):
    enabled: bool



class BulkAction(BaseModel):
    ids: List[str]
    action: Literal["verify", "reject"]


def now_iso():
    return datetime.now(timezone.utc).isoformat()


async def get_config():
    cfg = await db.config.find_one({"_id": "config"})
    if not cfg:
        cfg = {"_id": "config", "active_session": 1}
        await db.config.insert_one(cfg)
    return cfg


async def is_coming_soon():
    cfg = await get_config()
    return bool(cfg.get("coming_soon", True))


async def taken_seats(session_id: int):
    """Occupied seats for a session, sourced from atomic seat_locks.
    Expired locks (unpaid > HOLD_MINUTES) are auto-released and their orders marked expired."""
    now = datetime.now(timezone.utc)
    expired = await db.seat_locks.find(
        {"session_id": session_id, "expires_at": {"$ne": None, "$lt": now}}, {"order_id": 1}
    ).to_list(2000)
    if expired:
        order_ids = list({e["order_id"] for e in expired})
        await db.seat_locks.delete_many(
            {"session_id": session_id, "expires_at": {"$ne": None, "$lt": now}}
        )
        await db.orders.update_many(
            {"id": {"$in": order_ids}, "status": "pending_payment"},
            {"$set": {"status": "expired"}},
        )
    locks = await db.seat_locks.find({"session_id": session_id}, {"seat": 1}).to_list(2000)
    return set(l["seat"] for l in locks)


def build_seat_map(taken: set):
    """Pilihan kursi bebas: semua kursi yang belum dipesan bisa dipilih."""
    rows = []
    for r in SEAT_LAYOUT:
        blocks = []
        for blk in r["blocks"]:
            block_seats = []
            for n in blk:
                l = f"{r['row']}{n}"
                if l in RESERVED_SEATS:
                    status = "reserved"
                elif l in taken:
                    status = "booked"
                else:
                    status = "available"
                block_seats.append({
                    "label": l,
                    "status": status,
                    "couple": l in COUPLE_PARTNER,
                    "disability": l in DISABILITY_SEATS,
                })
            blocks.append(block_seats)
        rows.append({"row": r["row"], "unlocked": True, "blocks": blocks})
    return rows


def validate_couple_pairs(seats: list):
    """Kursi operator tidak bisa dipesan; kursi couple (pink) wajib sepasang."""
    seat_set = set(seats)
    for seat in seats:
        if seat in RESERVED_SEATS:
            raise HTTPException(
                status_code=400,
                detail=f"Kursi {seat} khusus operator dan tidak bisa dipesan",
            )
        partner = COUPLE_PARTNER.get(seat)
        if partner and partner not in seat_set:
            raise HTTPException(
                status_code=400,
                detail=f"Kursi {seat} adalah kursi Sweetbox — wajib dipesan sepasang dengan {partner}",
            )


async def session_status(session_id: int, open_sessions: set):
    taken = await taken_seats(session_id)
    is_full = len(taken) >= SEATS_PER_SESSION
    if session_id in open_sessions:
        status = "full" if is_full else "open"
    else:
        status = "closed"
    return status, len(taken)


async def get_open_sessions():
    """Sesi dibuka MANUAL oleh Super Admin. Default: semua tutup."""
    cfg = await get_config()
    return set(cfg.get("open_sessions", []))


async def get_walkin_sessions():
    """Sesi yang boleh dijual PANITIA di lokasi (walk-in). Terpisah dari sesi umum/online."""
    cfg = await get_config()
    return set(cfg.get("walkin_sessions", []))


def mask_name(name: str) -> str:
    parts = (name or "").strip().split()
    out = []
    for w in parts:
        if len(w) <= 2:
            out.append(w[0] + "*")
        else:
            out.append(w[0] + "*" * max(1, len(w) - 2) + w[-1])
    return " ".join(out)



async def gen_order_no():
    for _ in range(80):
        n = secrets.randbelow(9000) + 1000
        exists = await db.orders.count_documents({"order_no": n}, limit=1)
        if not exists:
            return n
    # fallback: sequential above current max
    last = await db.orders.find({"order_no": {"$exists": True}}).sort("order_no", -1).limit(1).to_list(1)
    return (last[0]["order_no"] + 1) if last else 1000


async def gen_unique_total(base: int):
    for pool in (list(range(11, 100)), list(range(100, 1000))):
        secrets.SystemRandom().shuffle(pool)
        for c in pool:
            exists = await db.orders.count_documents(
                {"total_amount": base + c, "status": {"$ne": "rejected"}}, limit=1
            )
            if not exists:
                return c, base + c
    raise HTTPException(status_code=409, detail="Tidak dapat membuat kode unik, coba lagi.")


def clean(o):
    o.pop("_id", None)
    return o


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode("utf-8"), hashed.encode("utf-8"))
    except Exception:
        return False


def create_token(user: dict) -> str:
    payload = {
        "sub": user["id"],
        "username": user["username"],
        "role": user["role"],
        "name": user.get("name", ""),
        "exp": datetime.now(timezone.utc) + timedelta(hours=TOKEN_TTL_HOURS),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)


def public_user(u: dict) -> dict:
    return {"id": u["id"], "username": u["username"], "name": u.get("name", ""),
            "role": u["role"], "role_label": ROLE_LABELS.get(u["role"], u["role"]),
            "can_delete": bool(u.get("can_delete", False)) or u["role"] == "superadmin",
            "created_at": u.get("created_at")}


async def log_activity(actor: dict, action: str, detail: str, target_id: str = None):
    await db.activity_logs.insert_one({
        "id": str(uuid.uuid4()),
        "actor_id": actor.get("id"),
        "actor_username": actor.get("username"),
        "actor_name": actor.get("name", ""),
        "action": action,
        "detail": detail,
        "target_id": target_id,
        "created_at": now_iso(),
    })


async def get_current_user(x_admin_token: Optional[str] = Header(None)):
    if not x_admin_token:
        raise HTTPException(status_code=401, detail="Tidak diizinkan")
    try:
        payload = jwt.decode(x_admin_token, JWT_SECRET, algorithms=[JWT_ALGO])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Sesi berakhir, silakan login ulang")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Tidak diizinkan")
    u = await db.admin_users.find_one({"id": payload.get("sub")})
    if not u:
        raise HTTPException(status_code=401, detail="Akun tidak ditemukan")
    return {"id": u["id"], "username": u["username"], "name": u.get("name", ""),
            "role": u["role"], "can_delete": bool(u.get("can_delete", False))}


def require_roles(*roles):
    async def dep(user: dict = Depends(get_current_user)):
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Akses tidak diizinkan untuk peran Anda")
        return user
    return dep


require_any = get_current_user
require_staff = require_roles("superadmin", "admin")
require_super = require_roles("superadmin")


# ---------------- Public endpoints ----------------
@api_router.get("/")
async def root():
    return {"message": "Nonton Bareng MBI API"}


@api_router.get("/event")
async def event_info():
    open_s = await get_open_sessions()
    walkin_s = await get_walkin_sessions()
    sessions = []
    for s in SESSIONS:
        status, count = await session_status(s["id"], open_s)
        sessions.append({**s, "status": status, "booked": count,
                         "capacity": SEATS_PER_SESSION,
                         "walkin_open": s["id"] in walkin_s})
    return {
        "title": EVENT_TITLE,
        "date": EVENT_DATE,
        "location": EVENT_LOCATION,
        "pricing": "donation",
        "reference_cost": REFERENCE_COST,
        "hold_minutes": HOLD_MINUTES,
        "sessions": sessions,
        "transfer": TRANSFER_INFO,
        "coming_soon": await is_coming_soon(),
    }


@api_router.get("/sessions/{session_id}/seats")
async def get_seats(session_id: int):
    if session_id not in [s["id"] for s in SESSIONS]:
        raise HTTPException(status_code=404, detail="Sesi tidak ditemukan")
    open_s = await get_open_sessions()
    status, count = await session_status(session_id, open_s)
    taken = await taken_seats(session_id)
    return {
        "session_id": session_id,
        "status": status,
        "rows": build_seat_map(taken),
        "booked": count,
        "capacity": SEATS_PER_SESSION,
        "couples": COUPLE_PARTNER,
    }


@api_router.post("/orders")
async def create_order(payload: OrderCreate):
    if await is_coming_soon():
        raise HTTPException(status_code=403, detail="Penjualan tiket belum dibuka")
    if not payload.name.strip() or not payload.phone.strip():
        raise HTTPException(status_code=400, detail="Nama dan nomor HP wajib diisi")
    if not payload.seats:
        raise HTTPException(status_code=400, detail="Pilih minimal 1 kursi")
    amount = int(payload.amount or 0)
    if amount <= 0:
        raise HTTPException(status_code=400, detail="Isi nominal dana sukarela terlebih dahulu")
    if amount > 100_000_000:
        raise HTTPException(status_code=400, detail="Nominal terlalu besar")

    open_s = await get_open_sessions()
    if payload.session_id not in open_s:
        raise HTTPException(status_code=400, detail="Sesi ini belum dibuka panitia")

    seats = list(dict.fromkeys(payload.seats))  # dedupe, keep order
    if len(seats) > MAX_SEATS_PER_ORDER:
        raise HTTPException(status_code=400, detail=f"Maksimal {MAX_SEATS_PER_ORDER} kursi per pemesanan online. Untuk lebih banyak, hubungi panitia.")
    for seat in seats:
        if seat not in ALL_SEAT_LABELS:
            raise HTTPException(status_code=400, detail=f"Kursi {seat} tidak valid")
    validate_couple_pairs(seats)
    for seat in seats:
        if seat in DISABILITY_SEATS:
            raise HTTPException(
                status_code=400,
                detail=f"Kursi {seat} khusus penyandang disabilitas — pembelian hanya di lokasi (walk-in)",
            )

    # Release any expired locks first so freed seats are claimable
    await taken_seats(payload.session_id)

    qty = len(seats)
    base = amount
    code, total = await gen_unique_total(base)
    order_no = await gen_order_no()
    order_id = str(uuid.uuid4())
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=HOLD_MINUTES)

    # Atomic seat claim: unique _id per (session, seat) prevents double booking
    claimed = []
    for seat in seats:
        try:
            await db.seat_locks.insert_one({
                "_id": f"{payload.session_id}:{seat}",
                "session_id": payload.session_id,
                "seat": seat,
                "order_id": order_id,
                "expires_at": expires_at,
            })
            claimed.append(seat)
        except DuplicateKeyError:
            if claimed:
                await db.seat_locks.delete_many(
                    {"_id": {"$in": [f"{payload.session_id}:{s}" for s in claimed]}}
                )
            raise HTTPException(status_code=409, detail=f"Kursi {seat} baru saja dipesan orang lain. Silakan pilih kursi lain.")

    order = {
        "id": order_id,
        "order_no": order_no,
        "name": payload.name.strip(),
        "phone": payload.phone.strip(),
        "session_id": payload.session_id,
        "seats": seats,
        "qty": qty,
        "base_amount": base,
        "unique_code": code,
        "total_amount": total,
        "payment_method": payload.payment_method,
        "status": "pending_payment",
        "proof_image": None,
        "checked_in": False,
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.orders.insert_one(order)
    return clean(dict(order))


@api_router.get("/orders/lookup")
async def lookup_orders(phone: str):
    if await is_coming_soon():
        raise HTTPException(status_code=403, detail="Penjualan tiket belum dibuka")
    p = phone.strip().replace(" ", "").replace("-", "")
    if len(p) < 6:
        raise HTTPException(status_code=400, detail="Masukkan nomor HP yang valid")
    # Filter di MongoDB berdasarkan digit nomor (abaikan spasi/strip) + batasi hasil,
    # supaya tidak menarik seluruh koleksi ke memori.
    digits = re.sub(r"\D", "", p)
    phone_pattern = r"\D*".join(re.escape(d) for d in digits)
    docs = await db.orders.find(
        {"phone": {"$regex": phone_pattern}, "status": {"$ne": "rejected"}, "deleted": {"$ne": True}},
        {"proof_image": 0}
    ).sort("created_at", -1).limit(200).to_list(200)
    result = []
    for o in docs:
        if o["phone"].replace(" ", "").replace("-", "") != p:
            continue
        # lazily expire unpaid too-old orders for accurate status
        session = next((s for s in SESSIONS if s["id"] == o["session_id"]), None)
        result.append({
            "id": o["id"], "order_no": o.get("order_no"), "name": mask_name(o["name"]), "phone": o["phone"],
            "session": session, "qty": o["qty"],
            "total_amount": o["total_amount"], "unique_code": o["unique_code"],
            "payment_method": o["payment_method"], "status": o["status"],
            "has_proof": bool(o.get("proof_image")), "created_at": o["created_at"],
        })
    return {"orders": result, "transfer": TRANSFER_INFO}


@api_router.get("/orders/{order_id}")
async def get_order(order_id: str):
    o = await db.orders.find_one({"id": order_id})
    if not o:
        raise HTTPException(status_code=404, detail="Pesanan tidak ditemukan")
    # lazily expire if unpaid and too old (also releases its seat locks)
    if o.get("status") == "pending_payment":
        created = datetime.fromisoformat(o["created_at"])
        if created < datetime.now(timezone.utc) - timedelta(minutes=HOLD_MINUTES):
            await db.orders.update_one({"id": order_id}, {"$set": {"status": "expired"}})
            await db.seat_locks.delete_many({"order_id": order_id})
            o["status"] = "expired"
    session = next((s for s in SESSIONS if s["id"] == o["session_id"]), None)
    o = clean(o)
    o["session"] = session
    o["transfer"] = TRANSFER_INFO
    return o


@api_router.post("/orders/{order_id}/proof")
async def upload_proof(order_id: str, payload: ProofUpload):
    o = await db.orders.find_one({"id": order_id})
    if not o:
        raise HTTPException(status_code=404, detail="Pesanan tidak ditemukan")
    if not payload.proof_image.startswith("data:image"):
        raise HTTPException(status_code=400, detail="File bukti harus berupa gambar")
    if len(payload.proof_image) > MAX_PROOF_BYTES:
        raise HTTPException(status_code=400, detail="Ukuran gambar bukti terlalu besar (maks ~6MB)")
    # Re-assert/keep seat reservation (handles late upload after auto-expire)
    for seat in o.get("seats", []):
        lid = f"{o['session_id']}:{seat}"
        existing = await db.seat_locks.find_one({"_id": lid})
        if existing:
            if existing["order_id"] == order_id:
                await db.seat_locks.update_one({"_id": lid}, {"$set": {"expires_at": None}})
            else:
                raise HTTPException(status_code=409, detail=f"Maaf, kursi {seat} sudah diambil peserta lain karena pembayaran melewati batas waktu. Silakan hubungi panitia.")
        else:
            await db.seat_locks.insert_one({
                "_id": lid, "session_id": o["session_id"], "seat": seat,
                "order_id": order_id, "expires_at": None,
            })
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {"proof_image": payload.proof_image, "status": "waiting_verification", "updated_at": now_iso()}},
    )
    o = await db.orders.find_one({"id": order_id})
    session = next((s for s in SESSIONS if s["id"] == o["session_id"]), None)
    o = clean(o)
    o["session"] = session
    o["transfer"] = TRANSFER_INFO
    return o


# ---------------- Admin endpoints ----------------
def client_ip(request: Request) -> str:
    xff = request.headers.get("x-forwarded-for", "")
    if xff:
        return xff.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


@api_router.post("/admin/login")
async def admin_login(payload: AdminLogin, request: Request):
    uname = payload.username.strip().lower()
    ident = f"{client_ip(request)}:{uname}"
    now = datetime.now(timezone.utc)
    rec = await db.login_attempts.find_one({"_id": ident})
    if rec and rec.get("locked_until") and rec["locked_until"].replace(tzinfo=timezone.utc) > now:
        wait = int((rec["locked_until"].replace(tzinfo=timezone.utc) - now).total_seconds() // 60) + 1
        raise HTTPException(status_code=429, detail=f"Terlalu banyak percobaan gagal. Coba lagi dalam {wait} menit.")
    u = await db.admin_users.find_one({"username": uname})
    if not u or not verify_password(payload.password, u["password_hash"]):
        fails = (rec.get("fails", 0) if rec else 0) + 1
        update = {"fails": fails, "updated_at": now}
        if fails >= LOGIN_MAX_FAILS:
            update["locked_until"] = now + timedelta(minutes=LOGIN_LOCK_MINUTES)
            update["fails"] = 0
        await db.login_attempts.update_one({"_id": ident}, {"$set": update}, upsert=True)
        raise HTTPException(status_code=401, detail="Username atau password salah")
    if rec:
        await db.login_attempts.delete_one({"_id": ident})
    token = create_token(u)
    return {"token": token, "user": public_user(u)}


@api_router.get("/admin/me")
async def admin_me(user: dict = Depends(get_current_user)):
    u = await db.admin_users.find_one({"id": user["id"]})
    return public_user(u)


@api_router.get("/admin/orders")
async def admin_orders(user: dict = Depends(require_staff), status: Optional[str] = None):
    match = {"deleted": {"$ne": True}}
    if status:
        match["status"] = status
    pipeline = []
    if match:
        pipeline.append({"$match": match})
    pipeline += [
        {"$addFields": {"has_proof": {"$and": [
            {"$ne": ["$proof_image", None]}, {"$ne": ["$proof_image", ""]},
        ]}}},
        {"$project": {"proof_image": 0}},
        {"$sort": {"created_at": -1}},
        {"$limit": 1000},
    ]
    orders = await db.orders.aggregate(pipeline).to_list(1000)
    result = []
    for o in orders:
        session = next((s for s in SESSIONS if s["id"] == o["session_id"]), None)
        o = clean(o)
        o["session"] = session
        result.append(o)
    return result


@api_router.get("/admin/orders/{order_id}/proof-image")
async def get_proof_image(order_id: str, user: dict = Depends(require_staff)):
    o = await db.orders.find_one({"id": order_id}, {"proof_image": 1})
    if not o or not o.get("proof_image"):
        raise HTTPException(status_code=404, detail="Bukti tidak ditemukan")
    return {"proof_image": o["proof_image"]}


@api_router.get("/admin/stats")
async def admin_stats(user: dict = Depends(require_staff)):
    pipeline = [{"$match": {"deleted": {"$ne": True}}}, {"$group": {
        "_id": "$status",
        "count": {"$sum": 1},
        "revenue": {"$sum": "$base_amount"},
        "tickets": {"$sum": "$qty"},
    }}]
    stats = {"total_orders": 0, "waiting_verification": 0, "verified": 0, "pending_payment": 0,
             "rejected": 0, "expired": 0, "revenue_verified": 0, "tickets_verified": 0}
    async for g in db.orders.aggregate(pipeline):
        st = g["_id"]
        stats["total_orders"] += g["count"]
        if st in stats:
            stats[st] += g["count"]
        if st == "verified":
            stats["revenue_verified"] += g.get("revenue", 0) or 0
            stats["tickets_verified"] += g.get("tickets", 0) or 0

    # Breakdown penjualan (hanya order terverifikasi): online vs walk-in per metode
    online = {"orders": 0, "tickets": 0, "revenue": 0}
    walkin = {"orders": 0, "tickets": 0, "revenue": 0,
              "cash": {"orders": 0, "tickets": 0, "revenue": 0},
              "qris": {"orders": 0, "tickets": 0, "revenue": 0},
              "transfer": {"orders": 0, "tickets": 0, "revenue": 0}}
    pipeline2 = [
        {"$match": {"status": "verified", "deleted": {"$ne": True}}},
        {"$group": {
            "_id": {"walkin": {"$ifNull": ["$walkin", False]}, "method": "$payment_method"},
            "orders": {"$sum": 1},
            "tickets": {"$sum": "$qty"},
            "revenue": {"$sum": "$base_amount"},
        }},
    ]
    async for g in db.orders.aggregate(pipeline2):
        is_walkin = bool(g["_id"].get("walkin"))
        method = g["_id"].get("method") or "transfer"
        o, t, r = g.get("orders", 0), g.get("tickets", 0), g.get("revenue", 0) or 0
        target = walkin if is_walkin else online
        target["orders"] += o
        target["tickets"] += t
        target["revenue"] += r
        if is_walkin and method in walkin:
            walkin[method]["orders"] += o
            walkin[method]["tickets"] += t
            walkin[method]["revenue"] += r
    stats["breakdown"] = {"online": online, "walkin": walkin}
    stats["cash_total"] = walkin["cash"]["revenue"]

    # Rekap dana per sesi (order terverifikasi)
    sess_agg = {}
    async for g in db.orders.aggregate([
        {"$match": {"status": "verified", "deleted": {"$ne": True}}},
        {"$group": {"_id": "$session_id", "orders": {"$sum": 1},
                    "tickets": {"$sum": "$qty"}, "revenue": {"$sum": "$base_amount"}}},
    ]):
        sess_agg[g["_id"]] = g
    stats["per_session"] = [{
        "id": s["id"], "name": s["name"], "time": s["time"],
        "orders": sess_agg.get(s["id"], {}).get("orders", 0),
        "tickets": sess_agg.get(s["id"], {}).get("tickets", 0),
        "revenue": sess_agg.get(s["id"], {}).get("revenue", 0) or 0,
    } for s in SESSIONS]

    # Rekap KAS CASH walk-in per sesi (jualan panitia di lokasi sebelum hari-H)
    cash_sess = {}
    async for g in db.orders.aggregate([
        {"$match": {"status": "verified", "walkin": True, "payment_method": "cash",
                    "deleted": {"$ne": True}}},
        {"$group": {"_id": "$session_id", "orders": {"$sum": 1},
                    "tickets": {"$sum": "$qty"}, "revenue": {"$sum": "$base_amount"}}},
    ]):
        cash_sess[g["_id"]] = g
    stats["cash_per_session"] = [{
        "id": s["id"], "name": s["name"], "time": s["time"],
        "orders": cash_sess.get(s["id"], {}).get("orders", 0),
        "tickets": cash_sess.get(s["id"], {}).get("tickets", 0),
        "revenue": cash_sess.get(s["id"], {}).get("revenue", 0) or 0,
    } for s in SESSIONS]

    # Riwayat pembelian per hari (WIB)
    wib = timezone(timedelta(hours=7))
    daily = {}
    async for o in db.orders.find(
        {"status": {"$nin": ["expired", "rejected"]}, "deleted": {"$ne": True}},
        {"created_at": 1, "qty": 1, "status": 1, "base_amount": 1},
    ):
        try:
            d = datetime.fromisoformat(o["created_at"]).astimezone(wib).strftime("%Y-%m-%d")
        except (ValueError, TypeError, KeyError):
            continue
        row = daily.setdefault(d, {"date": d, "orders": 0, "tickets": 0,
                                    "tickets_verified": 0, "revenue_verified": 0})
        row["orders"] += 1
        row["tickets"] += o.get("qty", 0) or 0
        if o.get("status") == "verified":
            row["tickets_verified"] += o.get("qty", 0) or 0
            row["revenue_verified"] += o.get("base_amount", 0) or 0
    stats["daily"] = [daily[k] for k in sorted(daily)]
    return stats


_MONTHS_ID = ["", "Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli",
              "Agustus", "September", "Oktober", "November", "Desember"]


async def _collect_recap_orders(date_from: str = None, date_to: str = None):
    """Semua pesanan TERVERIFIKASI (umum online + panitia walk-in), diperkaya untuk rekap.
    Opsional filter rentang tanggal WIB (YYYY-MM-DD, inklusif)."""
    wib = timezone(timedelta(hours=7))
    docs = await db.orders.find(
        {"status": "verified", "deleted": {"$ne": True}},
        {"proof_image": 0},
    ).sort("created_at", 1).to_list(8000)
    rows = []
    for o in docs:
        try:
            d = datetime.fromisoformat(o.get("created_at", "")).astimezone(wib)
        except (ValueError, TypeError):
            continue
        key = d.strftime("%Y-%m-%d")
        if date_from and key < date_from:
            continue
        if date_to and key > date_to:
            continue
        session = next((s for s in SESSIONS if s["id"] == o["session_id"]), None)
        is_walkin = bool(o.get("walkin"))
        is_vip = bool(o.get("vip"))
        if is_vip:
            channel, channel_label, seller, location = "vip", "VIP (Undangan)", "VIP (Undangan)", "VIP"
        elif is_walkin:
            channel, channel_label, seller = "panitia", "Panitia (Lokasi)", (o.get("sold_by") or "-")
            location = o.get("location") or "(tanpa lokasi)"
        else:
            channel, channel_label, seller, location = "umum", "Umum (Online)", "Umum (Online)", "Online"
        rows.append({
            "date": d.strftime("%Y-%m-%d"),
            "date_label": f"{d.day} {_MONTHS_ID[d.month]} {d.year}",
            "time": d.strftime("%H:%M"),
            "created_at": o.get("created_at"),
            "order_no": o.get("order_no"), "name": o.get("name"), "phone": o.get("phone", ""),
            "channel": channel,
            "channel_label": channel_label,
            "seller": seller,
            "location": location,
            "method": o.get("payment_method"),
            "tickets": o.get("qty", 0) or 0,
            "amount": o.get("total_amount", 0) or 0,
            "seats": o.get("seats", []),
            "session_id": o["session_id"],
            "session_name": session["name"] if session else str(o["session_id"]),
            "verified_by": o.get("verified_by", ""),
        })
    return rows


def _blank_agg():
    return {"cash": 0, "qris": 0, "transfer": 0, "amount": 0, "tickets": 0, "orders": 0,
            "umum_amount": 0, "panitia_amount": 0}


def _add_agg(acc, r):
    m = r.get("method", "cash")
    amt = r.get("amount", 0) or 0
    if m in ("cash", "qris", "transfer"):
        acc[m] += amt
    acc["amount"] += amt
    acc["tickets"] += r.get("tickets", 0) or 0
    acc["orders"] += 1
    if r.get("channel") == "panitia":
        acc["panitia_amount"] += amt
    else:
        acc["umum_amount"] += amt


@api_router.get("/admin/bendahara")
async def bendahara_recap(user: dict = Depends(require_staff),
                          date_from: str = Query(None, alias="from"),
                          date_to: str = Query(None, alias="to")):
    """Rekap keuangan LENGKAP: umum (online) + panitia (lokasi), per tanggal, petugas & lokasi & metode."""
    rows = await _collect_recap_orders(date_from, date_to)
    grand = _blank_agg()
    days = {}
    for r in rows:
        _add_agg(grand, r)
        day = days.setdefault(r["date"], {
            "date": r["date"], "date_label": r["date_label"],
            "total": _blank_agg(), "sellers": {}, "locations": {},
        })
        _add_agg(day["total"], r)
        _add_agg(day["sellers"].setdefault(r["seller"], _blank_agg()), r)
        _add_agg(day["locations"].setdefault(r["location"], _blank_agg()), r)

    by_date = []
    for key in sorted(days, reverse=True):
        day = days[key]
        by_date.append({
            "date": day["date"], "date_label": day["date_label"], "total": day["total"],
            "by_seller": [{"seller": k, **v} for k, v in sorted(day["sellers"].items())],
            "by_location": [{"location": k, **v} for k, v in sorted(day["locations"].items())],
        })
    # daftar transaksi (flat, terbaru dulu) untuk tabel yang bisa di-sort di frontend
    orders = sorted(rows, key=lambda r: r.get("created_at") or "", reverse=True)
    return {"grand_total": grand, "by_date": by_date, "orders": orders}


@api_router.get("/admin/bendahara/export")
async def export_bendahara(_: bool = Depends(require_staff),
                           date_from: str = Query(None, alias="from"),
                           date_to: str = Query(None, alias="to")):
    rows = await _collect_recap_orders(date_from, date_to)
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="255E33")

    def style_header(ws):
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Sheet 1: Semua Transaksi (detail)
    ws = wb.active
    ws.title = "Semua Transaksi"
    ws.append(["Tanggal", "Jam", "No. Order", "Nama Pembeli", "No HP", "Kanal", "Petugas/Sumber",
               "Lokasi", "Sesi", "Kursi", "Jml Tiket", "Metode", "Nominal (Rp)"])
    style_header(ws)
    method_label = {"cash": "Cash", "qris": "QRIS", "transfer": "Transfer"}
    for r in rows:
        ws.append([
            r["date"], r["time"], r["order_no"], r["name"], r["phone"], r["channel_label"],
            r["seller"], r["location"], r["session_name"], ", ".join(r["seats"]),
            r["tickets"], method_label.get(r["method"], r["method"]), r["amount"],
        ])
    for idx, w in enumerate([12, 7, 9, 22, 15, 16, 18, 20, 10, 16, 9, 10, 15], 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

    # Sheet 2: Ringkasan per Tanggal
    ws2 = wb.create_sheet("Ringkasan per Tanggal")
    ws2.append(["Tanggal", "Transaksi", "Tiket", "Cash", "QRIS", "Transfer",
                "Umum (Online)", "Panitia (Lokasi)", "TOTAL"])
    style_header(ws2)
    days = {}
    grand = _blank_agg()
    for r in rows:
        _add_agg(grand, r)
        _add_agg(days.setdefault(r["date"], _blank_agg()), r)
    for key in sorted(days):
        a = days[key]
        ws2.append([key, a["orders"], a["tickets"], a["cash"], a["qris"], a["transfer"],
                    a["umum_amount"], a["panitia_amount"], a["amount"]])
    ws2.append(["TOTAL", grand["orders"], grand["tickets"], grand["cash"], grand["qris"],
                grand["transfer"], grand["umum_amount"], grand["panitia_amount"], grand["amount"]])
    for cell in ws2[ws2.max_row]:
        cell.font = Font(bold=True)
    for idx, w in enumerate([12, 10, 8, 14, 14, 14, 16, 16, 16], 1):
        ws2.column_dimensions[ws2.cell(row=1, column=idx).column_letter].width = w

    # Sheet 3: Ringkasan per Petugas/Sumber
    ws3 = wb.create_sheet("Per Petugas")
    ws3.append(["Petugas/Sumber", "Transaksi", "Tiket", "Cash", "QRIS", "Transfer", "TOTAL"])
    style_header(ws3)
    sellers = {}
    for r in rows:
        _add_agg(sellers.setdefault(r["seller"], _blank_agg()), r)
    for k in sorted(sellers):
        a = sellers[k]
        ws3.append([k, a["orders"], a["tickets"], a["cash"], a["qris"], a["transfer"], a["amount"]])
    for idx, w in enumerate([20, 10, 8, 14, 14, 14, 16], 1):
        ws3.column_dimensions[ws3.cell(row=1, column=idx).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"rekap-bendahara-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@api_router.get("/admin/masterlist/export")
async def export_masterlist(_: bool = Depends(require_staff), type: str = Query("umum")):
    """Export Masterlist pembelian per tabel: type=umum (online+panitia) atau type=vip."""
    kind = (type or "umum").lower()
    if kind not in ("umum", "vip"):
        raise HTTPException(status_code=400, detail="type harus 'umum' atau 'vip'")
    rows = await _collect_recap_orders()
    if kind == "vip":
        rows = [r for r in rows if r["channel"] == "vip"]
    else:
        rows = [r for r in rows if r["channel"] != "vip"]
    rows.sort(key=lambda r: (r["session_id"], r["name"] or ""))

    wb = Workbook()
    ws = wb.active
    ws.title = "Tamu VIP" if kind == "vip" else "Pembeli Umum"
    header_fill = PatternFill("solid", fgColor="7A241F" if kind == "vip" else "255E33")
    if kind == "vip":
        ws.append(["No", "Nama", "No HP", "No. Order", "Sesi", "Kursi", "Jml Tiket"])
        widths = [5, 24, 16, 10, 12, 18, 9]
    else:
        ws.append(["No", "Nama", "No HP", "No. Order", "Sesi", "Kursi", "Jml Tiket", "Kanal", "Nominal (Rp)"])
        widths = [5, 24, 16, 10, 12, 18, 9, 16, 15]
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for i, r in enumerate(rows, 1):
        if kind == "vip":
            ws.append([i, r["name"], r["phone"], r["order_no"], r["session_name"],
                       ", ".join(r["seats"]), r["tickets"]])
        else:
            ws.append([i, r["name"], r["phone"], r["order_no"], r["session_name"],
                       ", ".join(r["seats"]), r["tickets"], r["channel_label"], r["amount"]])
    if kind != "vip" and rows:
        total = sum(r["amount"] for r in rows)
        ws.append(["", "TOTAL", "", "", "", "", sum(r["tickets"] for r in rows), "", total])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    label = "vip" if kind == "vip" else "umum"
    fname = f"masterlist-{label}-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )



@api_router.post("/admin/orders/{order_id}/verify")
async def verify_order(order_id: str, payload: VerifyPayload = VerifyPayload(), user: dict = Depends(require_staff)):
    o = await db.orders.find_one({"id": order_id})
    if not o:
        raise HTTPException(status_code=404, detail="Pesanan tidak ditemukan")
    actor = user.get("name") or user.get("username")
    fields = {"status": "verified", "verified_by": actor, "updated_at": now_iso()}
    note = f"Verifikasi pembayaran #{o.get('order_no')} — {o.get('name')}"
    if payload.amount is not None and payload.amount > 0 and payload.amount != o.get("total_amount"):
        old_total = o.get("total_amount")
        fields.update({"total_amount": payload.amount, "base_amount": payload.amount,
                       "amount_adjusted": True, "original_total": old_total})
        note += f" — nominal disesuaikan Rp{old_total:,} → Rp{payload.amount:,}".replace(",", ".")
    await db.orders.update_one({"id": order_id}, {"$set": fields})
    await log_activity(user, "verify", note, order_id)
    return clean(await db.orders.find_one({"id": order_id}))


@api_router.post("/admin/orders/{order_id}/reject")
async def reject_order(order_id: str, user: dict = Depends(require_staff)):
    o = await db.orders.find_one({"id": order_id})
    if not o:
        raise HTTPException(status_code=404, detail="Pesanan tidak ditemukan")
    actor = user.get("name") or user.get("username")
    await db.orders.update_one({"id": order_id}, {"$set": {
        "status": "rejected", "verified_by": actor, "updated_at": now_iso()}})
    await db.seat_locks.delete_many({"order_id": order_id})  # free the seats
    await log_activity(user, "reject", f"Tolak pesanan #{o.get('order_no')} — {o.get('name')}", order_id)
    return clean(await db.orders.find_one({"id": order_id}))


@api_router.post("/admin/orders/{order_id}/set-amount")
async def set_order_amount(order_id: str, payload: VerifyPayload, user: dict = Depends(require_staff)):
    o = await db.orders.find_one({"id": order_id})
    if not o:
        raise HTTPException(status_code=404, detail="Pesanan tidak ditemukan")
    if not payload.amount or payload.amount <= 0:
        raise HTTPException(status_code=400, detail="Nominal tidak valid")
    old = o.get("total_amount")
    await db.orders.update_one({"id": order_id}, {"$set": {
        "total_amount": payload.amount, "base_amount": payload.amount,
        "amount_adjusted": True, "original_total": o.get("original_total", old),
        "updated_at": now_iso()}})
    await log_activity(user, "verify",
                       f"Edit nominal #{o.get('order_no')} — {o.get('name')}: Rp{old:,} → Rp{payload.amount:,}".replace(",", "."),
                       order_id)
    return clean(await db.orders.find_one({"id": order_id}))


@api_router.delete("/admin/orders/{order_id}")
async def delete_order(order_id: str, user: dict = Depends(get_current_user)):
    if not (user["role"] == "superadmin" or user.get("can_delete")):
        raise HTTPException(status_code=403, detail="Akun Anda tidak diizinkan menghapus data")
    o = await db.orders.find_one({"id": order_id})
    if not o or o.get("deleted"):
        raise HTTPException(status_code=404, detail="Pesanan tidak ditemukan")
    actor = user.get("name") or user.get("username")
    await db.orders.update_one({"id": order_id}, {"$set": {
        "deleted": True, "deleted_at": now_iso(), "deleted_by": actor, "updated_at": now_iso()}})
    await db.seat_locks.delete_many({"order_id": order_id})  # free the seats
    await log_activity(user, "delete", f"Hapus pesanan #{o.get('order_no')} — {o.get('name')} ({o.get('phone')})", order_id)
    return {"deleted": True, "id": order_id}


@api_router.get("/admin/orders/deleted")
async def list_deleted_orders(user: dict = Depends(require_super)):
    docs = await db.orders.find({"deleted": True}, {"proof_image": 0}).sort("deleted_at", -1).to_list(1000)
    result = []
    for o in docs:
        session = next((s for s in SESSIONS if s["id"] == o["session_id"]), None)
        o = clean(o)
        o["session"] = session
        o["has_proof"] = False
        result.append(o)
    return result


@api_router.post("/admin/orders/{order_id}/restore")
async def restore_order(order_id: str, user: dict = Depends(require_super)):
    o = await db.orders.find_one({"id": order_id})
    if not o or not o.get("deleted"):
        raise HTTPException(status_code=404, detail="Pesanan terhapus tidak ditemukan")
    # Re-claim seats only for orders that still hold seats (active statuses)
    if o.get("status") in ("pending_payment", "waiting_verification", "verified"):
        await taken_seats(o["session_id"])  # release expired first
        claimed = []
        for seat in o.get("seats", []):
            lid = f"{o['session_id']}:{seat}"
            try:
                await db.seat_locks.insert_one({
                    "_id": lid, "session_id": o["session_id"], "seat": seat,
                    "order_id": order_id, "expires_at": None,
                })
                claimed.append(seat)
            except DuplicateKeyError:
                if claimed:
                    await db.seat_locks.delete_many({"_id": {"$in": [f"{o['session_id']}:{s}" for s in claimed]}})
                raise HTTPException(status_code=409, detail=f"Kursi {seat} sudah dipesan peserta lain — pesanan ini tidak bisa dipulihkan.")
    await db.orders.update_one({"id": order_id},
                               {"$set": {"deleted": False, "updated_at": now_iso()},
                                "$unset": {"deleted_at": "", "deleted_by": ""}})
    await log_activity(user, "restore", f"Pulihkan pesanan #{o.get('order_no')} — {o.get('name')}", order_id)
    return clean(await db.orders.find_one({"id": order_id}))


@api_router.post("/admin/orders/bulk")
async def bulk_action(payload: BulkAction, user: dict = Depends(require_staff)):
    if not payload.ids:
        raise HTTPException(status_code=400, detail="Tidak ada pesanan dipilih")
    if payload.action == "verify":
        await db.orders.update_many(
            {"id": {"$in": payload.ids}},
            {"$set": {"status": "verified", "updated_at": now_iso()}},
        )
    else:
        await db.orders.update_many(
            {"id": {"$in": payload.ids}},
            {"$set": {"status": "rejected", "updated_at": now_iso()}},
        )
        await db.seat_locks.delete_many({"order_id": {"$in": payload.ids}})
    act_label = "Verifikasi massal" if payload.action == "verify" else "Tolak massal"
    await log_activity(user, f"bulk_{payload.action}", f"{act_label} {len(payload.ids)} pesanan")
    return {"updated": len(payload.ids), "action": payload.action}


@api_router.post("/admin/orders/{order_id}/wa-sent")
async def mark_wa_sent(order_id: str, user: dict = Depends(require_staff)):
    r = await db.orders.update_one(
        {"id": order_id},
        {"$set": {"wa_sent": True, "wa_sent_at": now_iso(), "updated_at": now_iso()}},
    )
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Pesanan tidak ditemukan")
    return clean(await db.orders.find_one({"id": order_id}))


@api_router.post("/admin/orders/{order_id}/checkin")
async def checkin_order(order_id: str, user: dict = Depends(require_any)):
    o = await db.orders.find_one({"id": order_id})
    if not o:
        raise HTTPException(status_code=404, detail="Pesanan tidak ditemukan")
    already = o.get("checked_in", False)
    update = {"checked_in": True, "checked_in_at": now_iso(), "updated_at": now_iso()}
    if not already:
        update["checked_in_by"] = user.get("name") or user.get("username")
        update["checked_in_by_username"] = user.get("username")
    await db.orders.update_one({"id": order_id}, {"$set": update})
    if not already:
        await log_activity(user, "checkin", f"Check-in peserta #{o.get('order_no')} — {o.get('name')}", order_id)
    return clean(await db.orders.find_one({"id": order_id}))


@api_router.post("/admin/walkin")
async def walkin_order(payload: WalkinCreate, user: dict = Depends(require_staff)):
    if not payload.name.strip():
        raise HTTPException(status_code=400, detail="Nama wajib diisi")
    if payload.session_id not in [s["id"] for s in SESSIONS]:
        raise HTTPException(status_code=400, detail="Sesi tidak valid")
    walkin_s = await get_walkin_sessions()
    if payload.session_id not in walkin_s:
        raise HTTPException(status_code=400, detail="Sesi ini belum dibuka untuk penjualan panitia di lokasi. Minta Super Admin membukanya.")
    proof = (payload.proof_image or "").strip() or None
    if payload.payment_method in ("qris", "transfer"):
        if not proof or not proof.startswith("data:image"):
            raise HTTPException(status_code=400, detail="Untuk QRIS/Transfer wajib upload foto bukti pembayaran.")
        if len(proof) > MAX_PROOF_BYTES:
            raise HTTPException(status_code=400, detail="Ukuran gambar bukti terlalu besar (maks ~6MB)")
    else:
        proof = None  # cash tidak perlu bukti
    location = (payload.location or "").strip()
    if not location:
        raise HTTPException(status_code=400, detail="Lokasi penjualan wajib diisi")
    seats = list(dict.fromkeys(payload.seats))  # dedupe, keep order
    if not seats:
        raise HTTPException(status_code=400, detail="Pilih minimal 1 kursi")
    for seat in seats:
        if seat not in ALL_SEAT_LABELS:
            raise HTTPException(status_code=400, detail=f"Kursi {seat} tidak valid")
    validate_couple_pairs(seats)

    await taken_seats(payload.session_id)  # release expired locks first

    order_id = str(uuid.uuid4())
    claimed = []
    for seat in seats:
        try:
            await db.seat_locks.insert_one({
                "_id": f"{payload.session_id}:{seat}", "session_id": payload.session_id,
                "seat": seat, "order_id": order_id, "expires_at": None,
            })
            claimed.append(seat)
        except DuplicateKeyError:
            if claimed:
                await db.seat_locks.delete_many({"_id": {"$in": [f"{payload.session_id}:{s}" for s in claimed]}})
            raise HTTPException(status_code=409, detail=f"Kursi {seat} baru saja terisi. Pilih kursi lain.")

    qty = len(claimed)
    order_no = await gen_order_no()
    base = int(payload.amount or 0)
    if base <= 0:
        await db.seat_locks.delete_many({"_id": {"$in": [f"{payload.session_id}:{s}" for s in claimed]}})
        raise HTTPException(status_code=400, detail="Isi nominal dana sukarela terlebih dahulu")
    if base > 100_000_000:
        await db.seat_locks.delete_many({"_id": {"$in": [f"{payload.session_id}:{s}" for s in claimed]}})
        raise HTTPException(status_code=400, detail="Nominal terlalu besar")
    if payload.payment_method == "cash":
        code, total = 0, base
    else:
        code, total = await gen_unique_total(base)
    actor = user.get("name") or user.get("username")
    order = {
        "id": order_id, "order_no": order_no,
        "name": payload.name.strip(), "phone": (payload.phone or "").strip(),
        "session_id": payload.session_id, "seats": claimed, "qty": qty,
        "base_amount": base, "unique_code": code, "total_amount": total,
        "payment_method": payload.payment_method, "status": "verified",
        "proof_image": proof, "checked_in": True, "checked_in_at": now_iso(),
        "checked_in_by": actor, "checked_in_by_username": user.get("username"),
        "verified_by": actor,
        "walkin": True, "sold_by": actor,
        "location": location,
        "created_at": now_iso(), "updated_at": now_iso(),
    }
    await db.orders.insert_one(order)
    await log_activity(user, "walkin",
                       f"Walk-in {qty} tiket #{order_no} — {payload.name.strip()} ({payload.payment_method.upper()}) @ {location}, kursi {', '.join(claimed)}",
                       order_id)
    return clean(dict(order))



@api_router.post("/admin/vip")
async def vip_order(payload: VIPCreate, user: dict = Depends(require_staff)):
    """Pesan tiket VIP (tamu undangan) — gratis, kunci kursi di semua sesi, check-in menyusul."""
    if not payload.name.strip():
        raise HTTPException(status_code=400, detail="Isi nama tamu VIP")
    if payload.session_id not in [s["id"] for s in SESSIONS]:
        raise HTTPException(status_code=400, detail="Sesi tidak valid")
    seats = list(dict.fromkeys(payload.seats))
    if not seats:
        raise HTTPException(status_code=400, detail="Pilih minimal 1 kursi")
    for seat in seats:
        if seat not in ALL_SEAT_LABELS:
            raise HTTPException(status_code=400, detail=f"Kursi {seat} tidak valid")
        if seat in RESERVED_SEATS:
            raise HTTPException(status_code=400, detail=f"Kursi {seat} khusus operator, tidak bisa dipilih")

    await taken_seats(payload.session_id)  # release expired locks first
    order_id = str(uuid.uuid4())
    claimed = []
    for seat in seats:
        try:
            await db.seat_locks.insert_one({
                "_id": f"{payload.session_id}:{seat}", "session_id": payload.session_id,
                "seat": seat, "order_id": order_id, "expires_at": None,
            })
            claimed.append(seat)
        except DuplicateKeyError:
            if claimed:
                await db.seat_locks.delete_many({"_id": {"$in": [f"{payload.session_id}:{s}" for s in claimed]}})
            raise HTTPException(status_code=409, detail=f"Kursi {seat} sudah terisi. Pilih kursi lain.")

    actor = user.get("name") or user.get("username")
    order_no = await gen_order_no()
    order = {
        "id": order_id, "order_no": order_no,
        "name": payload.name.strip(), "phone": "",
        "session_id": payload.session_id, "seats": claimed, "qty": len(claimed),
        "base_amount": 0, "unique_code": 0, "total_amount": 0,
        "payment_method": "vip", "status": "verified",
        "proof_image": None, "checked_in": False,
        "verified_by": actor, "vip": True, "sold_by": actor,
        "note": (payload.note or "").strip(),
        "created_at": now_iso(), "updated_at": now_iso(),
    }
    await db.orders.insert_one(order)
    await log_activity(user, "vip",
                       f"Tiket VIP {len(claimed)} kursi #{order_no} — {payload.name.strip()}, kursi {', '.join(claimed)}",
                       order_id)
    return clean(dict(order))


@api_router.post("/admin/sessions/toggle")
async def toggle_session_open(payload: SessionToggle, user: dict = Depends(require_roles("superadmin"))):
    if payload.session_id not in [s["id"] for s in SESSIONS]:
        raise HTTPException(status_code=400, detail="Sesi tidak valid")
    key = "walkin_sessions" if payload.target == "walkin" else "open_sessions"
    label = "Panitia (Lokasi)" if payload.target == "walkin" else "Umum (Online)"
    cfg = await get_config()
    open_s = set(cfg.get(key, []))
    if payload.open:
        open_s.add(payload.session_id)
    else:
        open_s.discard(payload.session_id)
    await db.config.update_one({"_id": "config"}, {"$set": {key: sorted(open_s)}}, upsert=True)
    await log_activity(user, "session_toggle",
                       f"{'Membuka' if payload.open else 'Menutup'} Sesi {payload.session_id} untuk {label}")
    return {key: sorted(open_s)}


@api_router.post("/admin/coming-soon")
async def set_coming_soon(payload: SetComingSoon, user: dict = Depends(require_roles("superadmin"))):
    await db.config.update_one({"_id": "config"}, {"$set": {"coming_soon": payload.enabled}}, upsert=True)
    await log_activity(user, "coming_soon",
                       "Mengaktifkan mode Coming Soon (penjualan ditutup)" if payload.enabled
                       else "Membuka penjualan tiket (Coming Soon dimatikan)")
    return {"coming_soon": payload.enabled}


@api_router.get("/admin/participants")
async def list_participants(user: dict = Depends(require_any)):
    docs = await db.orders.find({"status": "verified", "deleted": {"$ne": True}}, {"proof_image": 0}).sort("created_at", -1).to_list(3000)
    result = []
    for o in docs:
        session = next((s for s in SESSIONS if s["id"] == o["session_id"]), None)
        result.append({
            "id": o["id"], "order_no": o.get("order_no"), "name": o["name"], "phone": o["phone"],
            "session": session, "seats": o["seats"], "qty": o["qty"],
            "checked_in": o.get("checked_in", False), "checked_in_at": o.get("checked_in_at"),
            "checked_in_by": o.get("checked_in_by"),
        })
    return result


@api_router.get("/admin/export")
async def export_orders(_: bool = Depends(require_staff)):
    status_label = {
        "pending_payment": "Belum Bayar",
        "waiting_verification": "Perlu Verifikasi",
        "verified": "Terverifikasi",
        "rejected": "Ditolak",
        "expired": "Kadaluarsa",
    }
    orders = await db.orders.find({"deleted": {"$ne": True}}, {"proof_image": 0}).sort([("session_id", 1), ("created_at", 1)]).to_list(5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Peserta"
    headers = ["No", "No. Order", "Kode Pesanan", "Nama", "No HP", "Sesi", "Jam", "Kursi",
               "Jml Tiket", "Nominal (Rp)", "Kode Unik", "Metode", "Kanal", "Dijual Oleh", "Lokasi Jual",
               "Diverifikasi Oleh", "Status", "Sudah Hadir", "Waktu Check-in", "Waktu Pesan"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for i, o in enumerate(orders, 1):
        session = next((s for s in SESSIONS if s["id"] == o["session_id"]), None)
        created = o.get("created_at", "")
        try:
            dt = datetime.fromisoformat(created)
            created_str = dt.strftime("%d/%m/%Y %H:%M")
        except Exception:
            created_str = created
        checkin_str = ""
        if o.get("checked_in_at"):
            try:
                checkin_str = datetime.fromisoformat(o["checked_in_at"]).strftime("%d/%m/%Y %H:%M")
            except Exception:
                checkin_str = o["checked_in_at"]
        method_label = {"qris": "QRIS", "transfer": "Transfer", "cash": "Cash"}.get(o.get("payment_method"), "Transfer")
        ws.append([
            i,
            o.get("order_no", ""),
            o["id"][:8].upper(),
            o["name"],
            o["phone"],
            session["name"] if session else o["session_id"],
            session["time"] if session else "",
            ", ".join(o.get("seats", [])),
            o.get("qty", 0),
            o.get("total_amount", 0),
            o.get("unique_code", ""),
            method_label,
            "Panitia (Lokasi)" if o.get("walkin") else "Umum (Online)",
            o.get("sold_by", "") if o.get("walkin") else "",
            o.get("location", "") if o.get("walkin") else "",
            o.get("verified_by", ""),
            status_label.get(o.get("status"), o.get("status")),
            "Ya" if o.get("checked_in") else "Belum",
            checkin_str,
            created_str,
        ])
    widths = [5, 9, 12, 24, 16, 10, 12, 18, 9, 14, 9, 10, 15, 16, 18, 16, 11, 18, 18]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w
    ws.freeze_panes = "A2"

    # Sheet Ringkasan: dana per sesi + pembelian per hari
    ws2 = wb.create_sheet("Ringkasan")
    bold = Font(bold=True)
    hdr_font = Font(bold=True, color="FFFFFF")

    ws2.append(["DANA TERKUMPUL PER SESI (terverifikasi)"])
    ws2["A1"].font = bold
    ws2.append(["Sesi", "Jam", "Pesanan", "Tiket", "Dana Terkumpul (Rp)"])
    for cell in ws2[2]:
        cell.font = hdr_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    verified = [o for o in orders if o.get("status") == "verified"]
    total_rev = 0
    for s in SESSIONS:
        so = [o for o in verified if o["session_id"] == s["id"]]
        rev = sum(o.get("base_amount", 0) or 0 for o in so)
        total_rev += rev
        ws2.append([s["name"], s["time"], len(so), sum(o.get("qty", 0) for o in so), rev])
    ws2.append(["TOTAL", "", len(verified), sum(o.get("qty", 0) for o in verified), total_rev])
    for cell in ws2[ws2.max_row]:
        cell.font = bold

    ws2.append([])
    r0 = ws2.max_row + 1
    ws2.append(["PEMBELIAN PER HARI (WIB, tanpa ditolak/kadaluarsa)"])
    ws2.cell(row=r0, column=1).font = bold
    ws2.append(["Tanggal", "Pembeli (pesanan)", "Tiket", "Tiket Terverifikasi", "Pendapatan Terverifikasi (Rp)"])
    for cell in ws2[ws2.max_row]:
        cell.font = hdr_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    wib = timezone(timedelta(hours=7))
    daily = {}
    for o in orders:
        if o.get("status") in ("expired", "rejected"):
            continue
        try:
            d = datetime.fromisoformat(o["created_at"]).astimezone(wib).strftime("%d/%m/%Y")
            key = datetime.fromisoformat(o["created_at"]).astimezone(wib).strftime("%Y-%m-%d")
        except (ValueError, TypeError, KeyError):
            continue
        row = daily.setdefault(key, {"label": d, "orders": 0, "tickets": 0, "tv": 0, "rev": 0})
        row["orders"] += 1
        row["tickets"] += o.get("qty", 0) or 0
        if o.get("status") == "verified":
            row["tv"] += o.get("qty", 0) or 0
            row["rev"] += o.get("base_amount", 0) or 0
    for k in sorted(daily):
        r = daily[k]
        ws2.append([r["label"], r["orders"], r["tickets"], r["tv"], r["rev"]])
    for idx, w in enumerate([14, 18, 10, 18, 26], 1):
        ws2.column_dimensions[ws2.cell(row=2, column=idx).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"peserta_nonton_mbi_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ---------------- User management (superadmin) ----------------
@api_router.get("/admin/users")
async def list_users(user: dict = Depends(require_super)):
    docs = await db.admin_users.find({}).sort("created_at", 1).to_list(200)
    return [public_user(u) for u in docs]


@api_router.post("/admin/users")
async def create_user(payload: UserCreate, user: dict = Depends(require_super)):
    uname = payload.username.strip().lower()
    if len(uname) < 3:
        raise HTTPException(status_code=400, detail="Username minimal 3 karakter")
    if len(payload.password) < 4:
        raise HTTPException(status_code=400, detail="Password minimal 4 karakter")
    if await db.admin_users.find_one({"username": uname}):
        raise HTTPException(status_code=409, detail="Username sudah dipakai")
    doc = {
        "id": str(uuid.uuid4()),
        "username": uname,
        "name": payload.name or uname,
        "role": payload.role,
        "can_delete": payload.can_delete,
        "password_hash": hash_password(payload.password),
        "created_at": now_iso(),
    }
    await db.admin_users.insert_one(doc)
    await log_activity(user, "user_create", f"Buat user '{uname}' ({ROLE_LABELS.get(payload.role, payload.role)})", doc["id"])
    return public_user(doc)


@api_router.post("/admin/users/{user_id}/permission")
async def set_user_permission(user_id: str, payload: UserPermission, user: dict = Depends(require_super)):
    target = await db.admin_users.find_one({"id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    await db.admin_users.update_one({"id": user_id}, {"$set": {"can_delete": payload.can_delete}})
    await log_activity(user, "user_permission",
                       f"{'Memberi' if payload.can_delete else 'Mencabut'} izin hapus data untuk '{target['username']}'", user_id)
    updated = await db.admin_users.find_one({"id": user_id})
    return public_user(updated)


@api_router.put("/admin/users/{user_id}")
async def update_user(user_id: str, payload: UserUpdate, user: dict = Depends(require_super)):
    target = await db.admin_users.find_one({"id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    updates = {}
    changes = []
    if payload.name is not None:
        nm = payload.name.strip()
        if len(nm) < 2:
            raise HTTPException(status_code=400, detail="Nama minimal 2 karakter")
        if nm != target.get("name"):
            updates["name"] = nm
            changes.append(f"nama → '{nm}'")
    if payload.role is not None and payload.role != target["role"]:
        # jaga jangan sampai super admin terakhir diturunkan
        if target["role"] == "superadmin" and payload.role != "superadmin":
            supers = await db.admin_users.count_documents({"role": "superadmin"})
            if supers <= 1:
                raise HTTPException(status_code=400, detail="Minimal harus ada 1 Super Admin")
        # cegah super admin menurunkan role dirinya sendiri (bisa terkunci)
        if user_id == user["id"] and payload.role != "superadmin":
            raise HTTPException(status_code=400, detail="Tidak bisa menurunkan role akun sendiri")
        updates["role"] = payload.role
        changes.append(f"peran → {ROLE_LABELS.get(payload.role, payload.role)}")
    if payload.can_delete is not None and bool(payload.can_delete) != bool(target.get("can_delete")):
        updates["can_delete"] = bool(payload.can_delete)
        changes.append(f"izin hapus → {'ya' if payload.can_delete else 'tidak'}")
    if not updates:
        return public_user(target)
    await db.admin_users.update_one({"id": user_id}, {"$set": updates})
    await log_activity(user, "user_update",
                       f"Ubah user '{target['username']}': {', '.join(changes)}", user_id)
    updated = await db.admin_users.find_one({"id": user_id})
    return public_user(updated)


@api_router.post("/admin/users/{user_id}/reset-password")
async def reset_user_password(user_id: str, payload: PasswordReset, user: dict = Depends(require_super)):
    target = await db.admin_users.find_one({"id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    pw = (payload.new_password or "").strip()
    if len(pw) < 6:
        raise HTTPException(status_code=400, detail="Password minimal 6 karakter")
    await db.admin_users.update_one({"id": user_id}, {"$set": {"password_hash": hash_password(pw)}})
    # bebaskan kunci brute-force untuk user ini (semua IP) agar bisa login ulang
    await db.login_attempts.delete_many({"_id": {"$regex": f":{target['username']}$"}})
    await log_activity(user, "reset_password", f"Reset password untuk '{target['username']}'", user_id)
    return {"reset": True, "id": user_id}


@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, user: dict = Depends(require_super)):
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="Tidak bisa menghapus akun sendiri")
    target = await db.admin_users.find_one({"id": user_id})
    if not target:
        raise HTTPException(status_code=404, detail="User tidak ditemukan")
    if target["role"] == "superadmin":
        supers = await db.admin_users.count_documents({"role": "superadmin"})
        if supers <= 1:
            raise HTTPException(status_code=400, detail="Minimal harus ada 1 Super Admin")
    await db.admin_users.delete_one({"id": user_id})
    await log_activity(user, "user_delete", f"Hapus user '{target['username']}'", user_id)
    return {"deleted": True, "id": user_id}


# ---------------- Activity logs (staff) ----------------
@api_router.get("/admin/logs")
async def list_logs(user: dict = Depends(require_staff), limit: int = 300):
    docs = await db.activity_logs.find({}).sort("created_at", -1).to_list(min(limit, 1000))
    return [clean(d) for d in docs]


ACTION_LABEL_ID = {
    "delete": "Hapus Order",
    "restore": "Pulihkan Order",
    "verify": "Verifikasi",
    "reject": "Tolak",
    "bulk_verify": "Verifikasi Massal",
    "bulk_reject": "Tolak Massal",
    "checkin": "Check-in",
    "user_create": "Buat User",
    "user_delete": "Hapus User",
    "user_update": "Ubah User",
    "user_permission": "Izin User",
    "reset_password": "Reset Password",
    "walkin": "Walk-in (Jual di Tempat)",
    "coming_soon": "Mode Coming Soon",
    "session_toggle": "Buka/Tutup Sesi",
}


@api_router.get("/admin/logs/export")
async def export_logs(_: bool = Depends(require_staff)):
    docs = await db.activity_logs.find({}).sort("created_at", -1).to_list(5000)
    wb = Workbook()
    ws = wb.active
    ws.title = "Log Aktivitas"
    headers = ["No", "Waktu (WIB)", "Petugas", "Username", "Aksi", "Detail"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for i, l in enumerate(docs, 1):
        waktu = l.get("created_at", "")
        try:
            dt = datetime.fromisoformat(waktu).astimezone(timezone(timedelta(hours=7)))
            waktu = dt.strftime("%d/%m/%Y %H:%M")
        except Exception:
            pass
        ws.append([
            i, waktu,
            l.get("actor_name") or l.get("actor_username", ""),
            l.get("actor_username", ""),
            ACTION_LABEL_ID.get(l.get("action"), l.get("action", "")),
            l.get("detail", ""),
        ])
    for idx, w in enumerate([5, 18, 20, 16, 16, 55], 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"log_aktivitas_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )




app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=False,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


@app.on_event("startup")
async def seed_admin_users():
    """Create default admin users on first run + unique index."""
    try:
        await db.admin_users.create_index("username", unique=True)
        count = await db.admin_users.count_documents({})
        if count == 0:
            default_pw = os.environ.get("ADMIN_PASSWORD", "admin123")
            defaults = [
                ("admin1", "Admin 1", "superadmin"),
                ("admin2", "Admin 2", "admin"),
                ("admin3", "Petugas Check-in", "checkin"),
            ]
            for uname, name, role in defaults:
                await db.admin_users.insert_one({
                    "id": str(uuid.uuid4()),
                    "username": uname,
                    "name": name,
                    "role": role,
                    "password_hash": hash_password(default_pw),
                    "created_at": now_iso(),
                })
            logger.info("default admin_users seeded")
    except Exception as e:
        logger.error(f"seed admin_users error: {e}")


@app.on_event("startup")
async def ensure_committee_users():
    """Idempotent: pastikan user panitia 'chelyn' ada (Admin + izin hapus data)."""
    try:
        if not await db.admin_users.find_one({"username": "chelyn"}):
            await db.admin_users.insert_one({
                "id": str(uuid.uuid4()),
                "username": "chelyn",
                "name": "Chelyn",
                "role": "admin",
                "can_delete": True,
                "password_hash": hash_password("Chelyn123456"),
                "created_at": now_iso(),
            })
            logger.info("committee user 'chelyn' ensured")
    except Exception as e:
        logger.error(f"ensure chelyn error: {e}")
    try:
        # auto-hapus catatan percobaan login lama (>1 hari)
        await db.login_attempts.create_index("updated_at", expireAfterSeconds=86400)
    except Exception as e:
        logger.error(f"login_attempts index error: {e}")


@app.on_event("startup")
async def backfill_seat_locks():
    """Ensure seat_locks exist for active (non-rejected/expired) orders.
    Runs on every startup and is idempotent (unique _id per session:seat)."""
    try:
        active_statuses = ["pending_payment", "waiting_verification", "verified"]
        async for o in db.orders.find({"status": {"$in": active_statuses}, "deleted": {"$ne": True}}):
            exp = None
            if o["status"] == "pending_payment":
                try:
                    exp = datetime.fromisoformat(o["created_at"]) + timedelta(minutes=HOLD_MINUTES)
                except Exception:
                    exp = None
            for seat in o.get("seats", []):
                try:
                    await db.seat_locks.insert_one({
                        "_id": f"{o['session_id']}:{seat}",
                        "session_id": o["session_id"],
                        "seat": seat,
                        "order_id": o["id"],
                        "expires_at": exp,
                    })
                except DuplicateKeyError:
                    pass
        logger.info("seat_locks backfill complete")
    except Exception as e:
        logger.error(f"seat_locks backfill error: {e}")
    # assign 4-digit order_no to any order missing it
    try:
        async for o in db.orders.find({"order_no": {"$exists": False}}):
            await db.orders.update_one({"id": o["id"]}, {"$set": {"order_no": await gen_order_no()}})
        logger.info("order_no backfill complete")
    except Exception as e:
        logger.error(f"order_no backfill error: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
