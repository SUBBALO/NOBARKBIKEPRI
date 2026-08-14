"""Iteration 10: 4 new sessions, bendahara (umum+panitia) + Excel export,
walk-in location required + proof required for QRIS/Transfer, tab removals impl on FE.
Backend only here."""
import io
import os
import time
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
API = f"{BASE_URL}/api"

SUPER = ("admin1", "admin123")
ADMIN = ("admin2", "admin123")
CHECKIN = ("admin3", "admin123")
CHELYN = ("chelyn", "Chelyn123456")


def login(u, p):
    r = requests.post(f"{API}/admin/login", json={"username": u, "password": p}, timeout=30)
    assert r.status_code == 200, f"login {u}: {r.status_code} {r.text}"
    return r.json()["token"]


@pytest.fixture(scope="module")
def tokens():
    return {
        "super": login(*SUPER),
        "admin": login(*ADMIN),
        "checkin": login(*CHECKIN),
        "chelyn": login(*CHELYN),
    }


def H(tok):
    return {"X-Admin-Token": tok}


# ---------- (A) SESSIONS ----------
def test_event_has_exactly_4_sessions_new_times():
    r = requests.get(f"{API}/event", timeout=30)
    assert r.status_code == 200
    sessions = r.json()["sessions"]
    assert len(sessions) == 4, f"Expected 4 sessions, got {len(sessions)}"
    expected = [
        (1, "09.30–11.30"),
        (2, "12.00–14.00"),
        (3, "14.30–16.30"),
        (4, "17.00–19.00"),
    ]
    for s, (eid, etime) in zip(sessions, expected):
        assert s["id"] == eid
        assert etime in s["time"], f"session {eid} time = {s['time']} vs expected {etime}"
        assert "walkin_open" in s and isinstance(s["walkin_open"], bool)
        assert "open" in s or "status" in s  # session status field exists


# ---------- (B) BENDAHARA ROLES ----------
def test_bendahara_permissions(tokens):
    r1 = requests.get(f"{API}/admin/bendahara", headers=H(tokens["super"]), timeout=30)
    assert r1.status_code == 200
    r2 = requests.get(f"{API}/admin/bendahara", headers=H(tokens["admin"]), timeout=30)
    assert r2.status_code == 200
    r3 = requests.get(f"{API}/admin/bendahara", headers=H(tokens["checkin"]), timeout=30)
    assert r3.status_code == 403


def test_bendahara_structure(tokens):
    r = requests.get(f"{API}/admin/bendahara", headers=H(tokens["super"]), timeout=30)
    assert r.status_code == 200
    data = r.json()
    assert set(data.keys()) >= {"grand_total", "by_date", "orders"}
    g = data["grand_total"]
    for k in ["cash", "qris", "transfer", "amount", "tickets", "orders",
              "umum_amount", "panitia_amount"]:
        assert k in g, f"missing key {k} in grand_total"
    # amount consistency
    assert g["amount"] == g["umum_amount"] + g["panitia_amount"], \
        f"amount {g['amount']} != umum {g['umum_amount']} + panitia {g['panitia_amount']}"
    assert g["amount"] == g["cash"] + g["qris"] + g["transfer"], \
        f"amount {g['amount']} != cash+qris+transfer {g['cash']+g['qris']+g['transfer']}"
    # by_date descending
    dates = [d["date"] for d in data["by_date"]]
    assert dates == sorted(dates, reverse=True)
    for d in data["by_date"]:
        assert "total" in d and "by_seller" in d and "by_location" in d
    # orders flat structure
    if data["orders"]:
        o = data["orders"][0]
        for k in ["channel", "seller", "location", "amount", "method", "order_no", "date"]:
            assert k in o


def test_bendahara_includes_umum_and_panitia_channels(tokens):
    """Check that if there are verified umum & panitia orders in DB, both appear."""
    r = requests.get(f"{API}/admin/bendahara", headers=H(tokens["super"]), timeout=30)
    data = r.json()
    channels = {o["channel"] for o in data["orders"]}
    # We can't force data to exist without polluting DB. Just assert channel values are valid.
    assert channels.issubset({"umum", "panitia"})
    # sellers/locations for umum should be 'Umum (Online)' / 'Online'
    for o in data["orders"]:
        if o["channel"] == "umum":
            assert o["seller"] == "Umum (Online)"
            assert o["location"] == "Online"


# ---------- (B2) EXPORT EXCEL ----------
def test_bendahara_export_xlsx(tokens):
    r = requests.get(f"{API}/admin/bendahara/export", headers=H(tokens["super"]), timeout=60)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "").lower() or \
           r.headers.get("content-type", "").endswith("sheet")
    wb = load_workbook(io.BytesIO(r.content))
    assert set(wb.sheetnames) == {"Semua Transaksi", "Ringkasan per Tanggal", "Per Petugas"}, \
        f"sheets = {wb.sheetnames}"


def test_bendahara_export_denied_for_checkin(tokens):
    r = requests.get(f"{API}/admin/bendahara/export", headers=H(tokens["checkin"]), timeout=30)
    assert r.status_code == 403


# ---------- (C) WALK-IN LOCATION + PROOF ----------
def _open_walkin(tokens, session_id, opened=True):
    r = requests.post(f"{API}/admin/sessions/toggle",
                      headers=H(tokens["super"]),
                      json={"session_id": session_id, "open": opened, "target": "walkin"},
                      timeout=30)
    assert r.status_code == 200, r.text


def test_walkin_requires_location_cash(tokens):
    _open_walkin(tokens, 4, True)
    try:
        # cash + no location -> 400 location error
        payload = {
            "session_id": 4, "name": "TEST_iter10_nolocation",
            "phone": "0800", "seats": ["L2"], "amount": 60000,
            "payment_method": "cash",
        }
        r = requests.post(f"{API}/admin/walkin", headers=H(tokens["super"]), json=payload, timeout=30)
        assert r.status_code == 400
        assert "Lokasi" in r.json().get("detail", "") or "lokasi" in r.json().get("detail", "").lower()
    finally:
        pass  # keep open for next tests, close at end


def test_walkin_qris_requires_proof(tokens):
    _open_walkin(tokens, 4, True)
    # qris with location but no proof -> 400
    payload = {
        "session_id": 4, "name": "TEST_iter10_noproof",
        "phone": "0800", "seats": ["L3"], "amount": 60000,
        "payment_method": "qris", "location": "Test Booth A",
    }
    r = requests.post(f"{API}/admin/walkin", headers=H(tokens["super"]), json=payload, timeout=30)
    assert r.status_code == 400
    assert "bukti" in r.json().get("detail", "").lower()


CREATED_ORDER_IDS = []


def test_walkin_cash_with_location_success(tokens):
    _open_walkin(tokens, 4, True)
    payload = {
        "session_id": 4, "name": "TEST_iter10_cash",
        "phone": "0801", "seats": ["M6"], "amount": 60000,
        "payment_method": "cash", "location": "Vihara TEST",
    }
    r = requests.post(f"{API}/admin/walkin", headers=H(tokens["super"]), json=payload, timeout=30)
    assert r.status_code == 200, r.text
    o = r.json()
    assert o["status"] == "verified"
    assert o["checked_in"] is True
    assert o["walkin"] is True
    assert o["location"] == "Vihara TEST"
    assert o.get("sold_by")
    CREATED_ORDER_IDS.append(o["id"])


TINY_PNG = (
    "data:image/png;base64,"
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNkYAAAAAYAAjCB0C8AAAAASUVORK5CYII="
)


def test_walkin_qris_with_proof_success(tokens):
    _open_walkin(tokens, 4, True)
    payload = {
        "session_id": 4, "name": "TEST_iter10_qris",
        "phone": "0802", "seats": ["M7"], "amount": 60000,
        "payment_method": "qris", "location": "Vihara TEST",
        "proof_image": TINY_PNG,
    }
    r = requests.post(f"{API}/admin/walkin", headers=H(tokens["super"]), json=payload, timeout=30)
    assert r.status_code == 200, r.text
    o = r.json()
    assert o["status"] == "verified"
    assert o["location"] == "Vihara TEST"
    CREATED_ORDER_IDS.append(o["id"])


def test_walkin_included_in_bendahara(tokens):
    """After creating walk-in orders, they should appear in bendahara panitia channel."""
    r = requests.get(f"{API}/admin/bendahara", headers=H(tokens["super"]), timeout=30)
    data = r.json()
    panitia_orders = [o for o in data["orders"] if o["channel"] == "panitia"]
    order_nos = {o["order_no"] for o in panitia_orders}
    # Get our created walk-in order_nos
    for oid in CREATED_ORDER_IDS:
        det = requests.get(f"{API}/admin/orders/{oid}", headers=H(tokens["super"]), timeout=15)
        if det.status_code == 200:
            assert det.json()["order_no"] in order_nos, f"order {det.json()['order_no']} missing"
    # Panitia amount must be > 0 now
    assert data["grand_total"]["panitia_amount"] >= 120000


# ---------- Regression: soft-delete + restore + can_delete ----------
def test_chelyn_can_soft_delete_and_super_restore(tokens):
    if not CREATED_ORDER_IDS:
        pytest.skip("no order created")
    oid = CREATED_ORDER_IDS[0]
    # chelyn deletes (should succeed - can_delete)
    r = requests.delete(f"{API}/admin/orders/{oid}", headers=H(tokens["chelyn"]), timeout=30)
    assert r.status_code in (200, 204), r.text
    # verify not in main list
    r2 = requests.get(f"{API}/admin/orders", headers=H(tokens["super"]), timeout=30)
    ids = [o["id"] for o in r2.json()]
    assert oid not in ids
    # deleted list contains it
    r3 = requests.get(f"{API}/admin/orders/deleted", headers=H(tokens["super"]), timeout=30)
    assert oid in [o["id"] for o in r3.json()]
    # restore
    r4 = requests.post(f"{API}/admin/orders/{oid}/restore", headers=H(tokens["super"]), timeout=30)
    assert r4.status_code == 200


def test_soft_deleted_not_in_bendahara(tokens):
    """Soft-delete a walk-in, ensure it is excluded from bendahara."""
    if len(CREATED_ORDER_IDS) < 2:
        pytest.skip()
    oid = CREATED_ORDER_IDS[1]
    r = requests.delete(f"{API}/admin/orders/{oid}", headers=H(tokens["super"]), timeout=30)
    assert r.status_code in (200, 204)
    b = requests.get(f"{API}/admin/bendahara", headers=H(tokens["super"]), timeout=30).json()
    ids_flat = {o["order_no"] for o in b["orders"]}
    # get its order_no from trash
    trash = requests.get(f"{API}/admin/orders/deleted", headers=H(tokens["super"]), timeout=30).json()
    on = next((t["order_no"] for t in trash if t["id"] == oid), None)
    if on:
        assert on not in ids_flat


# ---------- Teardown: clean up test orders ----------
def test_zzz_cleanup(tokens):
    for oid in CREATED_ORDER_IDS:
        requests.delete(f"{API}/admin/orders/{oid}", headers=H(tokens["super"]), timeout=15)
    # close walkin session 4
    requests.post(f"{API}/admin/sessions/toggle",
                  headers=H(tokens["super"]),
                  json={"session_id": 4, "open": False, "target": "walkin"},
                  timeout=15)
